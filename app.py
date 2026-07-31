from datetime import datetime
import io
import os
import openpyxl
import pandas as pd
import streamlit as st

# Sayfa Yapılandırması
st.set_page_config(
    page_title="Yalçın Market - Teknik Servis Takip",
    page_icon="🛠️",
    layout="wide",
)

# Yüklenen görseller ve belgeler için klasör oluşturma
UPLOADS_DIR = "uploads"
os.makedirs(UPLOADS_DIR, exist_ok=True)

file_path = "Yalcin_Market_Gelismis_Teknik_Servis_Takip_Sistemi.xlsx"


def ensure_excel_exists():
    if os.path.exists(file_path):
        try:
            df_test = pd.read_excel(
                file_path, sheet_name="Arıza Takip Listesi", skiprows=16
            )
            df_test = df_test.dropna(subset=["Sorun / Arıza Açıklaması"])
            if df_test.empty:
                os.remove(file_path)
        except:
            try:
                os.remove(file_path)
            except:
                pass

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Arıza Takip Listesi"
        ws.cell(
            row=1,
            column=2,
            value="YALÇIN MARKET - DETAYLI TEKNİK ARIZA VE MÜDAHALE TAKİP FORMU",
        )

        headers = [
            None,
            "Sıra",
            "Bildirim Tarih/Saat",
            "Şube Adı",
            "Sorun / Arıza Açıklaması",
            "Kategori",
            "Öncelik",
            "Durum",
            "Atanan Personel",
            "SLA Durumu",
            "Çözüm Süresi / Açıklama",
            "Fotoğraf / Belge",
            "Yön. Onay",
        ]
        for col_idx, header in enumerate(headers):
            if header:
                ws.cell(row=17, column=col_idx, value=header)

        ornekler = [
            [
                None,
                1,
                "2026-07-29 08:30",
                "Merkez Şube",
                "Şarküteri reyonu soğutma derecesi yükseliyor",
                "Soğutma / Soğuk Zincir",
                "Kritik",
                "Tamamlandı",
                "Ali Usta",
                "Zamanında",
                "Kompresör gazı yenilendi, test edildi.",
                "Yok",
                "Onaylandı",
            ],
            [
                None,
                2,
                "2026-07-29 09:15",
                "Şube 02 - Bahçelievler",
                "Kasa 2 barkod okuyucu temassızlık",
                "Kasa & IT Donanım",
                "Normal",
                "Devam Ediyor",
                "Caner Bey",
                "Devam Ediyor",
                "Kablo değişimi yapılacak.",
                "Yok",
                "Bekliyor",
            ],
            [
                None,
                3,
                "2026-07-29 10:00",
                "Merkez Şube",
                "Depo koridor aydınlatma armatürü arızalı",
                "Elektrik & Aydınlatma",
                "Düşük",
                "Tamamlandı",
                "Ahmet Usta",
                "Zamanında",
                "LED ampul değiştirildi.",
                "Yok",
                "Onaylandı",
            ],
        ]

        for idx, satir in enumerate(ornekler):
            for col_idx, val in enumerate(satir):
                if val is not None:
                    ws.cell(row=18 + idx, column=col_idx, value=val)

        wb.save(file_path)


ensure_excel_exists()


def calculate_sla(df):
    """Otomatik SLA Takip Mantığı:
    Devam eden arızalarda bildirim tarihinden itibaren geçen süreyi hesaplar.
    Kritik arızalar 24 saati geçmişse SLA 'Gecikti' olarak güncellenir.
    """
    if df.empty:
        return df

    now = datetime.now()
    for idx, row in df.iterrows():
        if row["Durum"] != "Tamamlandı":
            try:
                tarih_str = str(row["Bildirim Tarih/Saat"])
                tarih_obj = datetime.strptime(tarih_str, "%Y-%m-%d %H:%M")
                gecen_saat = (now - tarih_obj).total_seconds() / 3600

                # Kritik arızalarda 24 saat, normalde 48 saat sınırı
                limit_saat = 24 if row["Öncelik"] == "Kritik" else 48

                if gecen_saat > limit_saat:
                    df.at[idx, "SLA Durumu"] = "Gecikti"
                else:
                    df.at[idx, "SLA Durumu"] = "Devam Ediyor"
            except Exception:
                pass
    return df


def load_data():
    try:
        df = pd.read_excel(file_path, sheet_name="Arıza Takip Listesi", skiprows=16)
        if "Sıra" not in df.columns and len(df.columns) > 1:
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        df = df.dropna(subset=["Sorun / Arıza Açıklaması"])
        df = calculate_sla(df)
        return df
    except Exception:
        return pd.DataFrame(
            columns=[
                "Sıra",
                "Bildirim Tarih/Saat",
                "Şube Adı",
                "Sorun / Arıza Açıklaması",
                "Kategori",
                "Öncelik",
                "Durum",
                "Atanan Personel",
                "SLA Durumu",
                "Çözüm Süresi / Açıklama",
                "Fotoğraf / Belge",
                "Yön. Onay",
            ]
        )


df_ariza = load_data()

# Başlık
st.title("🛠️ Yalçın Market Teknik Servis ve Arıza Takip Sistemi")

# SLA Acil Uyarı Bandı
gecikenler = (
    df_ariza[
        (df_ariza["SLA Durumu"] == "Gecikti")
        & (df_ariza["Durum"] != "Tamamlandı")
    ]
    if not df_ariza.empty
    else pd.DataFrame()
)
if not gecikenler.empty:
    st.error(
        f"🚨 **ACİL DİKKAT:** Müdahale süresi (SLA) aşılmış **{len(gecikenler)} adet** çözülmeyen arıza bulunmaktadır!"
    )

st.markdown("---")

# Üst Özet Kartları (KPIs)
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("Toplam Arıza Kaydı", len(df_ariza))
with col2:
    devam_eden = (
        len(df_ariza[df_ariza["Durum"] == "Devam Ediyor"])
        if not df_ariza.empty
        else 0
    )
    st.metric("Devam Eden (Açık)", devam_eden)
with col3:
    kritik = (
        len(df_ariza[df_ariza["Öncelik"] == "Kritik"])
        if not df_ariza.empty
        else 0
    )
    st.metric("Kritik Öncelikli", kritik)
with col4:
    geciken_sayi = len(gecikenler)
    st.metric(
        "SLA Süresi Geciken",
        geciken_sayi,
        delta="- Acil Müdahale" if geciken_sayi > 0 else "Normal",
        delta_color="inverse",
    )

st.markdown("---")

# --- YÖNETİCİ DASHBOARD VE GRAFİKLER ---
st.markdown("### 📊 Dashboard ve Analiz Grafikleri")
if not df_ariza.empty:
    g_col1, g_col2 = st.columns(2)
    with g_col1:
        st.markdown("#### Şubelere Göre Arıza Dağılımı")
        st.bar_chart(df_ariza["Şube Adı"].value_counts())
    with g_col2:
        st.markdown("#### Arıza Kategorileri Dağılımı")
        st.bar_chart(df_ariza["Kategori"].value_counts(), color="#7c3aed")

st.markdown("---")

# --- DETAYLI FİLTRELEME VE ARAMA PANELİ ---
st.markdown("### 🔍 Detaylı Arama ve Filtreleme")
f_col1, f_col2, f_col3, f_col4 = st.columns(4)

with f_col1:
    arama_metni = st.text_input(
        "🔎 Arıza/Açıklama Ara", placeholder="Örn: soğutma, kasa..."
    )
with f_col2:
    subeler = (
        ["Tümü"] + list(df_ariza["Şube Adı"].dropna().unique())
        if not df_ariza.empty
        else ["Tümü"]
    )
    secilen_sube = st.selectbox("Şube Filtresi", subeler)
with f_col3:
    kategoriler = (
        ["Tümü"] + list(df_ariza["Kategori"].dropna().unique())
        if not df_ariza.empty
        else ["Tümü"]
    )
    secilen_kategori = st.selectbox("Kategori Filtresi", kategoriler)
with f_col4:
    durumlar = ["Tümü", "Devam Ediyor", "Tamamlandı", "İptal Edildi"]
    secilen_durum = st.selectbox("Durum Filtresi", durumlar)

# Filtrelerin Uygulanması
df_goster = df_ariza.copy()
if arama_metni:
    df_goster = df_goster[
        df_goster["Sorun / Arıza Açıklaması"]
        .astype(str)
        .str.contains(arama_metni, case=False, na=False)
    ]
if secilen_sube != "Tümü":
    df_goster = df_goster[df_goster["Şube Adı"] == secilen_sube]
if secilen_kategori != "Tümü":
    df_goster = df_goster[df_goster["Kategori"] == secilen_kategori]
if secilen_durum != "Tümü":
    df_goster = df_goster[df_goster["Durum"] == secilen_durum]

# --- ARIZA LİSTESİ TABLOSU VE İNDİRME BUTONLARI ---
col_t1, col_t2 = st.columns([3, 1])
with col_t1:
    st.markdown(f"### 📋 Arıza Takip Listesi ({len(df_goster)} Kayıt Found)")
with col_t2:
    # Excel Rapor İndirme Butonu
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_goster.to_excel(
            writer, sheet_name="Ariza_Raporu", index=False, startrow=2
        )
    excel_data = output.getvalue()

    st.download_button(
        label="📥 Excel Raporunu İndir",
        data=excel_data,
        file_name=f"Yalcin_Market_Ariza_Raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if not df_goster.empty:
    st.dataframe(df_goster, use_container_width=True)
else:
    st.info("Filtreleme kriterlerine uygun arıza kaydı bulunamadı.")

# --- YÜKLENEN GÖRSELLERİ/BELGELERİ İNCELEME ---
if not df_ariza.empty and "Fotoğraf / Belge" in df_ariza.columns:
    gorselli_arizalar = df_ariza[
        (df_ariza["Fotoğraf / Belge"].notna())
        & (df_ariza["Fotoğraf / Belge"] != "Yok")
    ]
    if not gorselli_arizalar.empty:
        with st.expander("📷 Yüklenmiş Arıza Görselleri / Belgeleri İncele"):
            g_secim = st.selectbox(
                "Görselini Görmek İstediğiniz Arıza Kaydı",
                gorselli_arizalar["Sıra"].astype(str)
                + " - "
                + gorselli_arizalar["Sorun / Arıza Açıklaması"],
            )
            if g_secim:
                secilen_id = g_secim.split(" - ")[0]
                dosya_adi = gorselli_arizalar[
                    gorselli_arizalar["Sıra"].astype(str) == secilen_id
                ]["Fotoğraf / Belge"].values[0]
                dosya_yolu = os.path.join(UPLOADS_DIR, str(dosya_adi))
                if os.path.exists(dosya_yolu):
                    st.image(
                        dosya_yolu,
                        caption=f"Arıza #{secilen_id} Kanıt Görseli",
                        width=400,
                    )
                else:
                    st.warning("Görsel dosyası sunucuda bulunamadı.")

st.markdown("---")

# --- ARIZA DURUMU VE ÇÖZÜM GÜNCELLEME FORMU ---
st.markdown("### ⚙️ Arıza Müdahale, Çözüm ve Görsel Ekleme")
with st.form("guncelleme_formu"):
    col_g1, col_g2, col_g3 = st.columns(3)

    with col_g1:
        secilen_sira = st.selectbox(
            "İşlem Yapılacak Arıza (Sıra No)",
            df_ariza["Sıra"].astype(str).tolist() if not df_ariza.empty else [],
        )
    with col_g2:
        yeni_durum = st.selectbox(
            "Yeni Durum", ["Devam Ediyor", "Tamamlandı", "İptal Edildi"]
        )
    with col_g3:
        atanan_personel = st.text_input(
            "Atanan Teknik Personel", value="Ali Usta"
        )

    cozum_aciklamasi = st.text_area(
        "Çözüm Açıklaması / Yapılan İşlem",
        placeholder="Örn: Parça değiştirildi, test edildi.",
    )
    yuklenen_dosya_cozum = st.file_uploader(
        "Servis Fişi / Çözüm Görseli Yükle (İsteğe Bağlı)",
        type=["png", "jpg", "jpeg", "pdf"],
    )

    guncelle_submit = st.form_submit_button("Arızayı Güncelle")

    if guncelle_submit:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb["Arıza Takip Listesi"]

            kaydedilen_dosya_adi = "Yok"
            if yuklenen_dosya_cozum is not None:
                kaydedilen_dosya_adi = (
                    f"cozum_{secilen_sira}_{yuklenen_dosya_cozum.name}"
                )
                with open(
                    os.path.join(UPLOADS_DIR, kaydedilen_dosya_adi), "wb"
                ) as f:
                    f.write(yuklenen_dosya_cozum.getbuffer())

            bulundu = False
            for row in range(18, ws.max_row + 1):
                sira_hucre = ws.cell(row=row, column=2).value
                if sira_hucre is not None and str(sira_hucre) == str(
                    secilen_sira
                ):
                    ws.cell(row=row, column=8, value=yeni_durum)
                    ws.cell(row=row, column=9, value=atanan_personel)
                    ws.cell(row=row, column=11, value=cozum_aciklamasi)
                    if kaydedilen_dosya_adi != "Yok":
                        ws.cell(row=row, column=12, value=kaydedilen_dosya_adi)
                    bulundu = True
                    break

            if bulundu:
                wb.save(file_path)
                st.success(
                    f"#{secilen_sira} numaralı arıza başarıyla güncellendi!"
                )
                st.rerun()
            else:
                st.error("Arıza kaydı Excel dosyasında bulunamadı.")
        except Exception as e:
            st.error(f"Güncelleme sırasında hata oluştu: {e}")

# --- YENİ ARIZA KAYDI FORMU (SİDEBAR) ---
st.sidebar.header("➕ Yeni Arıza Bildirimi")
with st.sidebar.form("ariza_form"):
    yeni_sube = st.selectbox(
        "Şube Adı",
        [
            "Merkez Şube",
            "Şube 02 - Bahçelievler",
            "Şube 03 - Meydan",
            "Şube 05 - Çarşı",
        ],
    )
    yeni_kategori = st.selectbox(
        "Kategori",
        [
            "Soğutma / Soğuk Zincir",
            "Kasa & IT Donanım",
            "Elektrik & Aydınlatma",
            "HVAC (Klima / Havalandırma)",
            "Mekanik & Raf / Kapı",
        ],
    )
    yeni_oncelik = st.selectbox("Öncelik", ["Kritik", "Normal", "Düşük"])
    yeni_aciklama = st.text_area("Arıza Açıklaması")
    yuklenen_dosya_ariza = st.sidebar.file_uploader(
        "Arıza Fotoğrafı Yükle", type=["png", "jpg", "jpeg", "pdf"]
    )

    submit = st.form_submit_button("Arıza Kaydı Oluştur")

    if submit:
        if not yeni_aciklama.strip():
            st.sidebar.error("Lütfen bir arıza açıklaması girin!")
        else:
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb["Arıza Takip Listesi"]

                gercek_son_satir = 17
                for r in range(18, ws.max_row + 2):
                    val = ws.cell(row=r, column=5).value
                    if val is not None and str(val).strip() != "":
                        gercek_son_satir = r

                yeni_hedef_satir = gercek_son_satir + 1
                yeni_sira = len(df_ariza) + 1
                simdiki_zaman = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")

                kaydedilen_dosya_adi = "Yok"
                if yuklenen_dosya_ariza is not None:
                    kaydedilen_dosya_adi = (
                        f"ariza_{yeni_sira}_{yuklenen_dosya_ariza.name}"
                    )
                    with open(
                        os.path.join(UPLOADS_DIR, kaydedilen_dosya_adi), "wb"
                    ) as f:
                        f.write(yuklenen_dosya_ariza.getbuffer())

                ws.cell(row=yeni_hedef_satir, column=2, value=yeni_sira)
                ws.cell(row=yeni_hedef_satir, column=3, value=simdiki_zaman)
                ws.cell(row=yeni_hedef_satir, column=4, value=yeni_sube)
                ws.cell(row=yeni_hedef_satir, column=5, value=yeni_aciklama)
                ws.cell(row=yeni_hedef_satir, column=6, value=yeni_kategori)
                ws.cell(row=yeni_hedef_satir, column=7, value=yeni_oncelik)
                ws.cell(row=yeni_hedef_satir, column=8, value="Devam Ediyor")
                ws.cell(row=yeni_hedef_satir, column=9, value="Atanmadı")
                ws.cell(row=yeni_hedef_satir, column=10, value="Devam Ediyor")
                ws.cell(row=yeni_hedef_satir, column=11, value="İşlem Bekliyor")
                ws.cell(
                    row=yeni_hedef_satir, column=12, value=kaydedilen_dosya_adi
                )
                ws.cell(row=yeni_hedef_satir, column=13, value="Bekliyor")

                wb.save(file_path)
                st.sidebar.success("Arıza kaydı fotoğrafla birlikte eklendi!")
                st.rerun()
            except Exception as e:
                st.sidebar.error(f"Kayıt eklenirken hata oluştu: {e}")
